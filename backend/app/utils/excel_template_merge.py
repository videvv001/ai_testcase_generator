"""
Merge test cases into an uploaded Excel template.
Matches exact template structure: Summary sheet unchanged, Test Cases sheet
has rows 1-2 as headers (with merges), row 3+ as data. Columns A-L (1-12).
"""
from __future__ import annotations

import re
import tempfile
from copy import copy
from pathlib import Path
from typing import Any, List, Mapping, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Template: columns A-L (1-12); row 1-2 headers, row 3+ data
MAX_DATA_COLS = 12
SHEET_NAME = "Test Cases"
SUMMARY_SHEET = "Summary"
MAX_TEMPLATE_SIZE_BYTES = 10 * 1024 * 1024  # 10MB


def format_test_steps(steps_input: str | List[str] | None) -> str:
    """
    Format test steps as enumerated list.
    Input: "Navigate to page | Click button | Verify result"
    Output: "1. Navigate to page\n2. Click button\n3. Verify result"
    Handles N/A, None; strips existing numbering.
    """
    if steps_input is None:
        return ""
    if isinstance(steps_input, list):
        parts = [str(s).strip() for s in steps_input if s]
    else:
        s = str(steps_input).strip()
        if not s or s in ("N/A", "None", ""):
            return ""
        parts = [p.strip() for p in s.split("|") if p.strip()]
    if not parts:
        return ""
    formatted = []
    for i, step in enumerate(parts, 1):
        step = step.strip()
        step = re.sub(r"^\d+[.)]\s*", "", step)  # strip existing numbering
        formatted.append(f"{i}. {step}")
    return "\n".join(formatted)


def _feature_prefix(feature_name: str) -> str:
    """
    Generate prefix for Test IDs based on feature name.
    Examples: "login page" -> "LOGIN", "helper management" -> "HM"
    """
    if not feature_name or not feature_name.strip():
        return "GEN"
    stop_words = {"page", "the", "a", "an"}
    words = [w for w in feature_name.lower().split() if w not in stop_words]
    if not words:
        return feature_name[:5].upper() if len(feature_name) >= 5 else feature_name.upper()
    if len(words) == 1:
        return words[0][:5].upper()
    return "".join(w[0].upper() for w in words[:2])


def _get_style_dict(ws: Worksheet, template_row: int) -> dict[int, dict[str, Any]]:
    """Copy font, fill, border, alignment, number_format from template row (cols 1-12)."""
    styles: dict[int, dict[str, Any]] = {}
    for col in range(1, MAX_DATA_COLS + 1):
        cell = ws.cell(row=template_row, column=col)
        styles[col] = {}
        if cell.has_style:
            styles[col] = {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
            }
        styles[col]["number_format"] = cell.number_format
    return styles


def _apply_style(cell, style: dict[str, Any]) -> None:
    if not style:
        return
    if "font" in style:
        cell.font = style["font"]
    if "fill" in style:
        cell.fill = style["fill"]
    if "border" in style:
        cell.border = style["border"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "number_format" in style:
        cell.number_format = style["number_format"]


def _copy_row(
    ws_src: Worksheet,
    row_src: int,
    ws_dest: Worksheet,
    row_dest: int,
    *,
    include_number_format: bool = True,
) -> None:
    """Copy one row (value + style) from source to destination sheet (cols 1-12)."""
    for col in range(1, MAX_DATA_COLS + 1):
        src_cell = ws_src.cell(row=row_src, column=col)
        dest_cell = ws_dest.cell(row=row_dest, column=col)
        dest_cell.value = src_cell.value
        if src_cell.has_style:
            style: dict[str, Any] = {
                "font": copy(src_cell.font),
                "fill": copy(src_cell.fill),
                "border": copy(src_cell.border),
                "alignment": copy(src_cell.alignment),
            }
            if include_number_format:
                style["number_format"] = src_cell.number_format
            _apply_style(dest_cell, style)


def _tc_value(tc: Mapping[str, Any], key: str, default: str = "") -> str:
    """Get string from dict with snake_case or camelCase key."""
    v = tc.get(key) or tc.get(key.replace("_", ""))
    if v is None and "_" in key:
        c = "".join(w.capitalize() if i else w.lower() for i, w in enumerate(key.split("_")))
        v = tc.get(c)
    return str(v).strip() if v is not None else default


def _write_data_row(
    ws: Worksheet,
    row: int,
    tc: Mapping[str, Any],
    feature_prefix: str,
    idx: int,
    template_styles: dict[int, dict[str, Any]],
    *,
    global_no: int | None = None,
) -> None:
    """Write one test case row (row 3+), columns A-L, and apply template formatting.
    idx = number within feature (for Test ID). global_no = optional sequential No. across all (column A).
    """
    scenario = _tc_value(tc, "test_scenario") or _tc_value(tc, "testScenario")
    description = _tc_value(tc, "test_description") or _tc_value(tc, "description")
    precondition = _tc_value(tc, "pre_condition") or _tc_value(tc, "precondition")
    test_data = _tc_value(tc, "test_data") or _tc_value(tc, "testData")
    steps_raw = tc.get("test_steps") or tc.get("testSteps")
    if isinstance(steps_raw, list):
        steps_str = format_test_steps(steps_raw)
    else:
        steps_str = format_test_steps(steps_raw if isinstance(steps_raw, str) else "")
    expected = _tc_value(tc, "expected_result") or _tc_value(tc, "expectedResult")

    test_id = f"TC_{feature_prefix}_{str(idx).zfill(3)}"
    no_value = global_no if global_no is not None else idx

    ws.cell(row=row, column=1, value=no_value)
    ws.cell(row=row, column=2, value=test_id)
    ws.cell(row=row, column=3, value=scenario)
    ws.cell(row=row, column=4, value=description)
    ws.cell(row=row, column=5, value=precondition)
    ws.cell(row=row, column=6, value=test_data)
    ws.cell(row=row, column=7, value=steps_str)
    ws.cell(row=row, column=8, value=expected)
    ws.cell(row=row, column=9, value="")
    ws.cell(row=row, column=10, value="Not Executed")
    ws.cell(row=row, column=11, value="")
    ws.cell(row=row, column=12, value="")

    for col in range(1, MAX_DATA_COLS + 1):
        _apply_style(ws.cell(row=row, column=col), template_styles.get(col, {}))


def merge_test_cases_to_excel(
    template_path: str | Path,
    test_cases: List[Mapping[str, Any]],
    feature_name: str,
) -> str:
    """
    Export test cases for a single feature to the template.
    - Summary sheet is left unchanged.
    - Test Cases sheet: rows 1-2 kept as-is, replace data from row 3+ with new test cases.
    - Preserve all formatting (font, fill, border, alignment, number_format).
    - Columns A-L: No., Test ID, Test Scenario, Test Description, Pre-condition, Test Data,
      Step (enumerated), Expected Result, Actual Result, Status, Comment, (empty).
    Returns path to saved temporary .xlsx file.
    """
    path = Path(template_path)
    if not path.exists() or path.stat().st_size > MAX_TEMPLATE_SIZE_BYTES:
        raise ValueError("Template file missing or exceeds size limit")

    wb = load_workbook(path, read_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Template must contain a sheet named '{SHEET_NAME}'")
    ws = wb[SHEET_NAME]

    header_end_row = 2
    data_start_row = 3
    template_styles = _get_style_dict(ws, data_start_row) if ws.max_row >= data_start_row else {}

    # Delete existing data rows only (keep headers rows 1-2)
    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - header_end_row)

    feature_prefix = _feature_prefix(feature_name or "Export")

    for idx, tc in enumerate(test_cases, 1):
        row = data_start_row + idx - 1
        _write_data_row(ws, row, tc, feature_prefix, idx, template_styles)

    tmp = tempfile.NamedTemporaryFile(
        prefix="export_test_cases_",
        suffix=".xlsx",
        delete=False,
    )
    out_path = Path(tmp.name)
    tmp.close()
    wb.save(str(out_path))
    return str(out_path)


def merge_all_features_to_excel(
    template_path: str | Path,
    features_data: List[Tuple[str, List[Mapping[str, Any]]]],
) -> str:
    """
    Export all features into the template's single "Test Cases" sheet.
    - Summary sheet is kept from template (unchanged).
    - Test Cases sheet is kept; only data rows (row 3+) are replaced.
    - All features' test cases are combined in order: Feature1's cases first, then Feature2's, etc.
    - Column A (No.) = global sequential number (1, 2, 3, ...).
    - Column B (Test ID) = per-feature (e.g. TC_FEAT1_001, ..., TC_FEAT2_001, ...).
    Returns path to saved temporary .xlsx file.
    """
    path = Path(template_path)
    if not path.exists() or path.stat().st_size > MAX_TEMPLATE_SIZE_BYTES:
        raise ValueError("Template file missing or exceeds size limit")

    wb = load_workbook(path, read_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Template must contain a sheet named '{SHEET_NAME}'")
    ws = wb[SHEET_NAME]

    header_end_row = 2
    data_start_row = 3
    template_styles = _get_style_dict(ws, data_start_row) if ws.max_row >= data_start_row else {}

    # Delete existing data rows only (keep headers rows 1-2)
    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - header_end_row)

    global_row = 1  # Sequential No. across all features (1, 2, 3, ...)
    current_data_row = data_start_row  # Next row to write (3, 4, 5, ...)

    for feature_name, test_cases in features_data:
        feature_prefix = _feature_prefix(feature_name or "Feature")
        for idx, tc in enumerate(test_cases, 1):
            _write_data_row(
                ws,
                current_data_row,
                tc,
                feature_prefix,
                idx,
                template_styles,
                global_no=global_row,
            )
            global_row += 1
            current_data_row += 1

    tmp = tempfile.NamedTemporaryFile(
        prefix="export_all_test_cases_",
        suffix=".xlsx",
        delete=False,
    )
    out_path = Path(tmp.name)
    tmp.close()
    wb.save(str(out_path))
    return str(out_path)
