from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from schemas.testcase import TestCase


def test_cases_to_excel(
    cases: Iterable[TestCase],
    *,
    prefix: str = "generated_test_cases_",
) -> str:
    """
    Convert generated test cases into an Excel file and return the file path.

    - One row per test case.
    - Bold header row.
    - Auto-adjusted column widths.
    - Saved as a temporary file suitable for FastAPI file download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Define columns (order matters for Excel output).
    headers: List[str] = [
        "Test Scenario",
        "Test Description",
        "Pre-condition",
        "Test Data",
        "Test Steps",
        "Expected Result",
    ]

    # Write header row with bold font.
    bold_font = Font(bold=True)
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font

    # Populate rows.
    for case in cases:
        steps_str = "\n".join(case.test_steps)

        ws.append(
            [
                case.test_scenario,
                case.test_description,
                case.pre_condition,
                case.test_data,
                steps_str,
                case.expected_result,
            ]
        )

    # Auto-adjust column widths based on maximum content length.
    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column  # 1-based index
        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
            except Exception:
                cell_value = ""
            if len(cell_value) > max_length:
                max_length = len(cell_value)

        # Add a little extra padding to the calculated width.
        adjusted_width = max_length + 2 if max_length > 0 else 10
        ws.column_dimensions[get_column_letter(column_index)].width = adjusted_width

    # Save to a temporary file and return its path.
    tmp = tempfile.NamedTemporaryFile(
        prefix=prefix,
        suffix=".xlsx",
        delete=False,
    )
    tmp_path = Path(tmp.name)
    tmp.close()  # We only need the path; openpyxl will write the file.

    wb.save(str(tmp_path))

    return str(tmp_path)
